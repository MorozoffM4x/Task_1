#Импорты

import ssl
import urllib.request
import requests
import urllib3
import shutil
import os
import xlrd
from xlwt import Worksheet
import xlwings as xw

from openpyxl import Workbook
from openpyxl import load_workbook
from lxml import html, etree
from urllib.request import urlopen
from datetime import timedelta, datetime

download_path = os.path.abspath('Download') + '\\'
done_path = os.path.abspath('Done') + '\\'
codes_con = ["TERAVATT", "SOLARGEN", "ABAKANKS", "KREMTECH", "BUGULSES", "OSPRJKOM", "SUNPROJT", "SUNPROD2", "EFFEKTDB", "ECOENRUS", "PROJECT5", "PROJECT6", "UGKTGKN8", "KREMTECH", "UGKTGKN8", "FORTNGE2", "STARPROJ"]

#Отключение ошибки сертификата и предупреждений

try:
   _create_unverified_https_context = ssl._create_unverified_context
except AttributeError:
    pass
else:
    ssl._create_default_https_context = _create_unverified_https_context

urllib3.disable_warnings()

#Даты

dt_t = datetime.today()
dt = dt_t.strftime("%Y%m%d")
period_day = 7 #Количество дней, за которое надо получить данные

# openpyxl, как оказалось, не поддерживает .xls

def write_row(date, rge, rows):
    file = load_workbook(os.path.abspath('ТГ Конкуренты - задание.xlsm'), read_only=False, keep_vba=True)
    sheet = file['Конкуренты']
    max_row = sheet.max_row

    for a in range(4, max_row):
        if sheet.cell(row=a, column=4).value == rge:
            sheet.cell(row=a, column=8, value=datetime.strptime(date, "%Y-%m-%d").date())
            col = 9
            for m in rows:
                sheet.cell(row=a, column=col, value=m)
                col += 1
    file.save(os.path.abspath('ТГ Конкуренты - задание.xlsm'))

# xlrd, как оказалось, не поддерживает .xlsx

def parse_files(dates, codes, filename):
    file = download_path + filename
    workbook = xlrd.open_workbook(file)
    sheet = workbook.sheet_by_index(0)

    for a in range(7, sheet.nrows - 1):
        rge = sheet.cell_value(a, 0)
        date = sheet.cell_value(2, 2)
        rows = sheet.row_values(a, 4, 179)
        write_row(date, rge, rows)


def download(dates):
    if not os.path.exists(download_path):
        os.makedirs(download_path)
    headers = {'Content-Type': 'text/html', }
    response = requests.get('https://www.atsenergo.ru/nreport?rname=carana_sell_units&region=eur&rdate=' + dates,
                            headers=headers, verify=False)
    html = response.text

    with open(download_path + 'atsenergo_' + dates + '.html', 'w') as f:
        f.write(html)

    local = 'file:///' + download_path + 'atsenergo_' + dates + '.html'
    response = urlopen(local)
    htmlparser = etree.HTMLParser()
    tree = etree.parse(response, htmlparser)

    files_list_lxml = tree.xpath('//div[@class = "reports_files"]/table/tr/td[2]/a/@href')
    text_list_lxml = tree.xpath('//div[@class = "reports_files"]/table/tr/td[2]/a/text()')

    lenght = len(files_list_lxml)
    for a in range(lenght - 1):
        codes = text_list_lxml[a][9: -19]

        if codes in codes_con:
            urllib.request.urlretrieve('https://www.atsenergo.ru/nreport' + files_list_lxml[a],
                                       download_path + text_list_lxml[a])
            parse_files(dates, codes, text_list_lxml[a])


# Main
for a in range(period_day):
    dt2 = dt_t - timedelta(a)
    date_today = dt2.strftime("%Y%m%d")

    download(date_today)
    shutil.rmtree(download_path)

    vba_book = xw.Book(os.path.abspath('ТГ Конкуренты - задание.xlsm'))
    vba_macro = vba_book.macro("copy_1")
    vba_macro()